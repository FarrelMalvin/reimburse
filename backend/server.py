from fastapi import FastAPI, APIRouter, HTTPException, Header, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, uuid
from pathlib import Path
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt as pyjwt
from fpdf import FPDF
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'kantorplus-secret-2024-production')
pwd_context = CryptContext(schemes=['bcrypt'], deprecated='auto')

# Status flow untuk Bon Sementara (Pengajuan Perjalanan Dinas)
# pending -> approved_atasan -> approved_hrga -> approved_direktur -> approved_finance
BON_STATUS_FLOW = {
    "pending": {"next": "approved_atasan", "approver": "atasan"},
    "approved_atasan": {"next": "approved_hrga", "approver": "hrga"},
    "approved_hrga": {"next": "approved_direktur", "approver": "direktur"},
    "approved_direktur": {"next": "approved_finance", "approver": "finance"},
    "approved_finance": {"next": None, "approver": None}
}

# Status flow untuk Realisasi Bon
# pending -> approved_hrga -> approved_direktur -> approved_finance
REALISASI_STATUS_FLOW = {
    "pending": {"next": "approved_hrga", "approver": "hrga"},
    "approved_hrga": {"next": "approved_direktur", "approver": "direktur"},
    "approved_direktur": {"next": "approved_finance", "approver": "finance"},
    "approved_finance": {"next": None, "approver": None}
}

app = FastAPI()
api_router = APIRouter(prefix="/api")

# --- Models ---
class LoginRequest(BaseModel):
    email: str
    password: str

class AkomodasiData(BaseModel):
    kota_tujuan: str = ""
    nama_hotel: str = ""
    check_in: str = ""
    check_out: str = ""
    harga_per_malam: float = 0
    pembayaran: str = "Head Office"

class TransportasiData(BaseModel):
    jenis: str = ""
    dari_kota: str = ""
    ke_kota: str = ""
    jam_berangkat: str = ""

class EstimasiItem(BaseModel):
    kategori: str = ""
    uraian: str = ""
    quantity: int = 1
    jumlah: float = 0

class BonSementaraCreate(BaseModel):
    nik: str = ""
    jabatan: str = ""
    wilayah: str = ""
    tujuan: str
    periode_mulai: str
    periode_selesai: str
    keperluan: str
    akomodasi: Optional[AkomodasiData] = None
    transportasi_berangkat: Optional[TransportasiData] = None
    transportasi_kembali: Optional[TransportasiData] = None
    estimasi_items: Optional[List[EstimasiItem]] = None
    jumlah: float
    foto: Optional[str] = None

class RealisasiItemData(BaseModel):
    tanggal: str
    uraian: str
    quantity: int = 1
    harga_per_unit: float = 0
    total: float = 0

class RealisasiBonCreate(BaseModel):
    bon_sementara_id: str
    periode: str
    items: List[RealisasiItemData]
    bukti_transfer: Optional[str] = None

class BonAction(BaseModel):
    reason: Optional[str] = None

class PengaduanCreate(BaseModel):
    judul: str
    deskripsi: str
    kategori: str

class CutiCreate(BaseModel):
    tanggal_mulai: str
    tanggal_selesai: str
    jenis: str
    alasan: str

class InventarisCreate(BaseModel):
    nama: str
    kategori: str
    jumlah: int
    kondisi: str
    lokasi: str

# --- Auth ---
def create_token(user):
    return pyjwt.encode({
        'id': user['id'], 'email': user['email'], 'role': user['role'], 'name': user['name'],
        'exp': datetime.now(timezone.utc) + timedelta(hours=24)
    }, JWT_SECRET, algorithm='HS256')

async def get_current_user(authorization: str = Header(None)):
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail='Token diperlukan')
    token = authorization.split(' ')[1]
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='Token kedaluwarsa')
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail='Token tidak valid')

@api_router.post("/auth/login")
async def login(data: LoginRequest):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not pwd_context.verify(data.password, user['password']):
        raise HTTPException(status_code=401, detail='Email atau password salah')
    token = create_token(user)
    return {"token": token, "user": {"id": user['id'], "name": user['name'], "email": user['email'], "role": user['role']}}

@api_router.get("/auth/me")
async def get_me(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    return {"id": user['id'], "name": user['name'], "email": user['email'], "role": user['role']}

# --- Seed ---
@api_router.post("/seed")
async def seed_data():
    count = await db.users.count_documents({})
    if count > 0:
        return {"message": "Data sudah ada"}
    users = [
        {'id': str(uuid.uuid4()), 'name': 'Ahmad Pegawai', 'email': 'pegawai@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'pegawai'},
        {'id': str(uuid.uuid4()), 'name': 'Budi Atasan', 'email': 'atasan@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'atasan'},
        {'id': str(uuid.uuid4()), 'name': 'Dewi HRGA', 'email': 'hrga@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'hrga'},
        {'id': str(uuid.uuid4()), 'name': 'Eko Direktur', 'email': 'direktur@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'direktur'},
        {'id': str(uuid.uuid4()), 'name': 'Citra Finance', 'email': 'finance@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'finance'},
    ]
    await db.users.insert_many(users)
    return {"message": "Seed berhasil"}

# --- Dashboard ---
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    r = user['role']
    if r == 'pegawai':
        total_bon = await db.bon_sementara.count_documents({"user_id": user['id']})
        total_realisasi = await db.realisasi_bon.count_documents({"user_id": user['id']})
        pengaduan = await db.pengaduan.count_documents({"user_id": user['id']})
        cuti = await db.cuti.count_documents({"user_id": user['id']})
    elif r == 'atasan':
        total_bon = await db.bon_sementara.count_documents({"status": "pending"})
        total_realisasi = 0  # Atasan tidak approve realisasi
        pengaduan = await db.pengaduan.count_documents({"status": "pending"})
        cuti = await db.cuti.count_documents({"status": "pending"})
    elif r == 'hrga':
        total_bon = await db.bon_sementara.count_documents({"status": "approved_atasan"})
        total_realisasi = await db.realisasi_bon.count_documents({"status": "pending"})
        pengaduan = await db.pengaduan.count_documents({})
        cuti = await db.cuti.count_documents({})
    elif r == 'direktur':
        total_bon = await db.bon_sementara.count_documents({"status": "approved_hrga"})
        total_realisasi = await db.realisasi_bon.count_documents({"status": "approved_hrga"})
        pengaduan = 0
        cuti = 0
    else:  # finance
        total_bon = await db.bon_sementara.count_documents({"status": "approved_direktur"})
        total_realisasi = await db.realisasi_bon.count_documents({"status": "approved_direktur"})
        pengaduan = await db.pengaduan.count_documents({})
        cuti = await db.cuti.count_documents({})
    inventaris = await db.inventaris.count_documents({})
    return {"total_bon": total_bon, "total_realisasi": total_realisasi, "pengaduan_aktif": pengaduan, "total_cuti": cuti, "total_inventaris": inventaris}

# --- Helper: Generate No Bon ---
async def generate_no_bon():
    now = datetime.now(timezone.utc)
    month = now.strftime("%m")
    year = now.strftime("%Y")
    prefix = f"BS/GA/{month}/{year}"
    count = await db.bon_sementara.count_documents({"no_bon": {"$regex": f"^{prefix}"}})
    return f"{prefix}/{count + 1:03d}"

# --- Bon Sementara ---
@api_router.get("/bon-sementara")
async def get_bon_sementara(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    role = user['role']
    if role == 'pegawai':
        items = await db.bon_sementara.find({"user_id": user['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'atasan':
        # Atasan sees pending (needs to approve) + already processed
        items = await db.bon_sementara.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'hrga':
        # HRGA sees approved_atasan (needs to approve) + already processed
        items = await db.bon_sementara.find({"status": {"$in": ["approved_atasan", "approved_hrga", "approved_direktur", "approved_finance", "declined"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'direktur':
        # Direktur sees approved_hrga (needs to approve) + already processed
        items = await db.bon_sementara.find({"status": {"$in": ["approved_hrga", "approved_direktur", "approved_finance", "declined"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    else:  # finance
        # Finance sees approved_direktur (needs to approve) + already processed
        items = await db.bon_sementara.find({"status": {"$in": ["approved_direktur", "approved_finance", "declined"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return items

@api_router.post("/bon-sementara")
async def create_bon_sementara(data: BonSementaraCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    if user['role'] != 'pegawai':
        raise HTTPException(status_code=403, detail='Hanya pegawai')
    no_bon = await generate_no_bon()
    # Estimasi items custom
    estimasi_items = [item.model_dump() for item in data.estimasi_items] if data.estimasi_items else []
    ako = data.akomodasi.model_dump() if data.akomodasi else {}
    tb = data.transportasi_berangkat.model_dump() if data.transportasi_berangkat else {}
    tk = data.transportasi_kembali.model_dump() if data.transportasi_kembali else {}
    doc = {
        "id": str(uuid.uuid4()), "no_bon": no_bon, "user_id": user['id'], "user_name": user['name'],
        "nik": data.nik, "jabatan": data.jabatan, "wilayah": data.wilayah,
        "tujuan": data.tujuan, "periode_mulai": data.periode_mulai, "periode_selesai": data.periode_selesai,
        "keperluan": data.keperluan, "akomodasi": ako, "transportasi_berangkat": tb, "transportasi_kembali": tk,
        "estimasi_items": estimasi_items, "jumlah": data.jumlah, "foto": data.foto,
        "status": "pending", "declined_by": None, "decline_reason": None,
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.bon_sementara.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.put("/bon-sementara/{bon_id}/approve")
async def approve_bon_sementara(bon_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    bon = await db.bon_sementara.find_one({"id": bon_id}, {"_id": 0})
    if not bon:
        raise HTTPException(status_code=404, detail='Tidak ditemukan')
    
    role = user['role']
    current_status = bon['status']
    
    # Check if user can approve based on status flow
    flow = BON_STATUS_FLOW.get(current_status)
    if not flow or flow['approver'] != role:
        raise HTTPException(status_code=400, detail=f'Tidak dapat disetujui. Status saat ini: {current_status}, Role: {role}')
    
    new_status = flow['next']
    
    # Track approval history
    approval_history = bon.get('approval_history', [])
    approval_history.append({
        "action": "approved",
        "by": user['name'],
        "role": role,
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    await db.bon_sementara.update_one(
        {"id": bon_id}, 
        {"$set": {
            "status": new_status, 
            "approval_history": approval_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notification message based on role
    status_labels = {
        "approved_atasan": "Atasan Departemen",
        "approved_hrga": "HRGA",
        "approved_direktur": "Direktur",
        "approved_finance": "Finance (Final)"
    }
    
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), 
        "user_id": bon['user_id'], 
        "message": f"Bon Sementara '{bon['no_bon']}' disetujui oleh {status_labels.get(new_status, role)}", 
        "type": "approved", 
        "bon_id": bon_id, 
        "is_read": False, 
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Disetujui", "status": new_status}

@api_router.put("/bon-sementara/{bon_id}/decline")
async def decline_bon_sementara(bon_id: str, data: BonAction, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    bon = await db.bon_sementara.find_one({"id": bon_id}, {"_id": 0})
    if not bon:
        raise HTTPException(status_code=404)
    
    role = user['role']
    current_status = bon['status']
    
    # Check if this role can decline based on current status
    flow = BON_STATUS_FLOW.get(current_status)
    if not flow or flow['approver'] != role:
        raise HTTPException(status_code=403, detail='Tidak dapat menolak pengajuan ini')
    
    # Track approval history
    approval_history = bon.get('approval_history', [])
    approval_history.append({
        "action": "declined",
        "by": user['name'],
        "role": role,
        "reason": data.reason or "Tidak ada alasan",
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    await db.bon_sementara.update_one(
        {"id": bon_id}, 
        {"$set": {
            "status": "declined", 
            "declined_by": role, 
            "decline_reason": data.reason or "Tidak ada alasan",
            "approval_history": approval_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    role_labels = {
        "atasan": "Atasan Departemen",
        "hrga": "HRGA",
        "direktur": "Direktur",
        "finance": "Finance"
    }
    
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), 
        "user_id": bon['user_id'], 
        "message": f"Bon Sementara '{bon['no_bon']}' ditolak oleh {role_labels.get(role, role)}: {data.reason or '-'}", 
        "type": "declined", 
        "bon_id": bon_id, 
        "is_read": False, 
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Ditolak", "status": "declined"}

@api_router.put("/bon-sementara/{bon_id}/resubmit")
async def resubmit_bon_sementara(bon_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    if user['role'] != 'pegawai':
        raise HTTPException(status_code=403)
    bon = await db.bon_sementara.find_one({"id": bon_id, "user_id": user['id']}, {"_id": 0})
    if not bon or bon['status'] != 'declined':
        raise HTTPException(status_code=400, detail='Tidak bisa diajukan ulang')
    await db.bon_sementara.update_one({"id": bon_id}, {"$set": {"status": "pending", "declined_by": None, "decline_reason": None, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Diajukan ulang"}

# PDF Form Perjalanan Dinas (untuk Pegawai - bisa download setelah disetujui Atasan)
@api_router.get("/bon-sementara/{bon_id}/pdf-perjalanan-dinas")
async def pdf_perjalanan_dinas(bon_id: str, authorization: str = Header(None)):
    await get_current_user(authorization)
    bon = await db.bon_sementara.find_one({"id": bon_id}, {"_id": 0})
    if not bon:
        raise HTTPException(status_code=404)
    # Bisa download setelah disetujui atasan (atau lebih)
    valid_statuses = ['approved_atasan', 'approved_hrga', 'approved_direktur', 'approved_finance']
    if bon['status'] not in valid_statuses:
        raise HTTPException(status_code=400, detail='Belum disetujui atasan')
    
    # Get approval dates from history
    approval_history = bon.get("approval_history", [])
    atasan_date = "-"
    hrga_date = "-"
    direktur_date = "-"
    for approval in approval_history:
        if approval.get("role") == "atasan" and approval.get("action") == "approved":
            atasan_date = approval.get("at", "-")[:10] if approval.get("at") else "-"
        elif approval.get("role") == "hrga" and approval.get("action") == "approved":
            hrga_date = approval.get("at", "-")[:10] if approval.get("at") else "-"
        elif approval.get("role") == "direktur" and approval.get("action") == "approved":
            direktur_date = approval.get("at", "-")[:10] if approval.get("at") else "-"
    
    pdf = FPDF()
    pdf.add_page()
    
    # Header
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 8, 'FORMULIR RENCANA PERJALANAN DINAS KARYAWAN', ln=True, align='C')
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(0, 5, f'No. Dokumen: {bon["no_bon"]}   Revisi: 00   Tgl Terbit: {bon["created_at"][:10]}', ln=True, align='C')
    pdf.ln(5)
    
    # Employee & Trip Details
    pdf.set_font('Helvetica', '', 9)
    
    def add_field(label, value, width1=35, width2=60):
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(width1, 6, label)
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(5, 6, ':')
        pdf.cell(width2, 6, str(value) if value else '-')
    
    # Row 1
    add_field('Nama', bon.get('user_name', '-'))
    add_field('Wilayah', bon.get('wilayah', '-'))
    pdf.ln()
    # Row 2
    add_field('Jabatan', bon.get('jabatan', '-'))
    add_field('NIK', bon.get('nik', '-'))
    pdf.ln()
    # Row 3
    add_field('Tujuan', bon.get('tujuan', '-'))
    add_field('Periode', f'{bon["periode_mulai"]} s/d {bon["periode_selesai"]}')
    pdf.ln()
    # Row 4
    add_field('Keperluan', bon.get('keperluan', '-'), 35, 140)
    pdf.ln(8)
    
    # AKOMODASI - HOTEL Section
    pdf.set_draw_color(0, 0, 0)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(0, 6, 'AKOMODASI - HOTEL', ln=True)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)
    
    ako = bon.get('akomodasi', {})
    add_field('Kota Tujuan', ako.get('kota_tujuan', '-'))
    add_field('Nama Hotel', ako.get('nama_hotel', '-'))
    pdf.ln()
    add_field('Check In', ako.get('check_in', '-'))
    add_field('Check Out', ako.get('check_out', '-'))
    pdf.ln()
    add_field('Harga/Menginap', f'Rp {ako.get("harga_per_malam", 0):,.0f}')
    add_field('Pembayaran', ako.get('pembayaran', 'Head Office'))
    pdf.ln(8)
    
    # TRANSPORTASI Section
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(0, 6, 'TRANSPORTASI', ln=True)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)
    
    tb = bon.get('transportasi_berangkat', {})
    tk = bon.get('transportasi_kembali', {})
    
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(0, 5, 'Keberangkatan:', ln=True)
    pdf.set_font('Helvetica', '', 9)
    add_field('Dari', tb.get('dari_kota', '-'))
    add_field('Ke', tb.get('ke_kota', '-'))
    pdf.ln()
    add_field('Transportasi', tb.get('jenis', '-'))
    add_field('Jam Berangkat', tb.get('jam_berangkat', '-'))
    pdf.ln(3)
    
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(0, 5, 'Kedatangan:', ln=True)
    pdf.set_font('Helvetica', '', 9)
    add_field('Dari', tk.get('dari_kota', '-'))
    add_field('Ke', tk.get('ke_kota', '-'))
    pdf.ln()
    add_field('Transportasi', tk.get('jenis', '-'))
    add_field('Jam', tk.get('jam_berangkat', '-'))
    pdf.ln(8)
    
    # ESTIMASI BIAYA Section
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(0, 6, 'ESTIMASI BIAYA PERJALANAN DINAS (KAS BON)', ln=True)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)
    
    estimasi_items = bon.get('estimasi_items', [])
    pdf.set_font('Helvetica', '', 9)
    total_est = 0
    for i, item in enumerate(estimasi_items):
        jumlah = item.get('jumlah', 0) or 0
        pdf.cell(10, 5, f'{i+1}.')
        pdf.cell(80, 5, item.get('uraian', '-'))
        pdf.cell(0, 5, f'Rp {jumlah:,.0f}', ln=True)
        total_est += jumlah
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(10, 5, '')
    pdf.cell(80, 5, 'TOTAL')
    pdf.cell(0, 5, f'Rp {total_est:,.0f}', ln=True)
    pdf.ln(8)
    
    # Approval Section
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(45, 5, 'Diajukan Oleh,', align='C')
    pdf.cell(45, 5, 'Diketahui Oleh,', align='C')
    pdf.cell(45, 5, 'Diperiksa Oleh,', align='C')
    pdf.cell(45, 5, 'Disetujui Oleh,', align='C', ln=True)
    
    pdf.cell(45, 5, 'Karyawan', align='C')
    pdf.cell(45, 5, 'Atasan Langsung', align='C')
    pdf.cell(45, 5, 'HR & GA Dept.', align='C')
    pdf.cell(45, 5, 'Direktur', align='C', ln=True)
    pdf.ln(12)
    
    # Get approver names from history
    atasan_name = "-"
    hrga_name_pdf = "-"
    direktur_name = "-"
    for approval in approval_history:
        if approval.get("action") == "approved":
            if approval.get("role") == "atasan":
                atasan_name = approval.get("by", "-")
            elif approval.get("role") == "hrga":
                hrga_name_pdf = approval.get("by", "-")
            elif approval.get("role") == "direktur":
                direktur_name = approval.get("by", "-")
    
    pdf.set_font('Helvetica', 'B', 8)
    pdf.cell(45, 5, bon.get('user_name', '-'), align='C')
    pdf.cell(45, 5, atasan_name, align='C')
    pdf.cell(45, 5, hrga_name_pdf, align='C')
    pdf.cell(45, 5, direktur_name, align='C', ln=True)
    
    pdf.set_font('Helvetica', '', 7)
    pdf.cell(45, 4, f'Tgl: {bon["created_at"][:10]}', align='C')
    pdf.cell(45, 4, f'Tgl: {atasan_date}', align='C')
    pdf.cell(45, 4, f'Tgl: {hrga_date}', align='C')
    pdf.cell(45, 4, f'Tgl: {direktur_date}', align='C')
    
    return Response(content=bytes(pdf.output()), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=RPD-{bon['no_bon'].replace('/', '-')}.pdf"})

# PDF Register Bon Sementara (untuk Pegawai - setelah Finance approve)
@api_router.get("/bon-sementara/{bon_id}/pdf")
async def pdf_bon_sementara(bon_id: str, authorization: str = Header(None)):
    await get_current_user(authorization)
    bon = await db.bon_sementara.find_one({"id": bon_id}, {"_id": 0})
    if not bon:
        raise HTTPException(status_code=404)
    if bon['status'] != 'approved_finance':
        raise HTTPException(status_code=400, detail='Belum disetujui sepenuhnya')
    
    # Get HRGA name from approval history
    hrga_name = "-"
    approval_history = bon.get("approval_history", [])
    for approval in approval_history:
        if approval.get("role") == "hrga" and approval.get("action") == "approved":
            hrga_name = approval.get("by", "-")
            break
    
    # Get all approver names
    finance_name = "-"
    for approval in approval_history:
        if approval.get("role") == "finance" and approval.get("action") == "approved":
            finance_name = approval.get("by", "-")
            break
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'REGISTER BON SEMENTARA', ln=True, align='C')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, f'No. Dokumen: {bon["no_bon"]}   Revisi: 00   Tgl Terbit: {bon["created_at"][:10]}', ln=True, align='C')
    pdf.ln(8)
    pdf.set_font('Helvetica', '', 10)
    fields = [
        ("Tgl. Pengajuan", bon["created_at"][:10]),
        ("No. Bon Sementara", bon["no_bon"]),
        ("Nama Pengaju", hrga_name),
        ("NIK", bon.get("nik", "-")),
        ("Jabatan", bon.get("jabatan", "-")),
        ("Wilayah", bon.get("wilayah", "-")),
        ("Tujuan", bon.get("tujuan", "-")),
        ("Periode", f'{bon["periode_mulai"]} s/d {bon["periode_selesai"]}'),
        ("Keperluan", bon["keperluan"]),
        ("Jumlah", f'Rp {bon["jumlah"]:,.0f}'),
    ]
    for label, val in fields:
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(45, 7, f'{label}')
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(5, 7, ':')
        pdf.cell(0, 7, str(val), ln=True)
    
    pdf.ln(15)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(60, 7, 'Yang Membawa,', align='C')
    pdf.cell(60, 7, 'Kasir,', align='C')
    pdf.cell(60, 7, 'Disetujui,', align='C', ln=True)
    pdf.ln(18)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(60, 7, hrga_name, align='C')
    pdf.cell(60, 7, finance_name, align='C')
    pdf.cell(60, 7, '________________', align='C')
    return Response(content=bytes(pdf.output()), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={bon['no_bon'].replace('/', '-')}.pdf"})

# --- Realisasi Bon ---
@api_router.get("/realisasi")
async def get_realisasi(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    role = user['role']
    if role == 'pegawai':
        items = await db.realisasi_bon.find({"user_id": user['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'atasan':
        # Atasan tidak terlibat dalam realisasi, tapi bisa lihat semua
        items = await db.realisasi_bon.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'hrga':
        # HRGA approve realisasi pertama kali (pending -> approved_hrga)
        items = await db.realisasi_bon.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif role == 'direktur':
        # Direktur approve setelah HRGA
        items = await db.realisasi_bon.find({"status": {"$in": ["approved_hrga", "approved_direktur", "approved_finance", "declined"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    else:  # finance
        # Finance approve terakhir
        items = await db.realisasi_bon.find({"status": {"$in": ["approved_direktur", "approved_finance", "declined"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return items

@api_router.get("/bon-sementara-approved")
async def get_approved_bons(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    bons = await db.bon_sementara.find({"user_id": user['id'], "status": "approved_finance"}, {"_id": 0, "foto": 0}).to_list(100)
    return bons

@api_router.post("/realisasi")
async def create_realisasi(data: RealisasiBonCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    if user['role'] != 'pegawai':
        raise HTTPException(status_code=403)
    bon = await db.bon_sementara.find_one({"id": data.bon_sementara_id, "user_id": user['id']}, {"_id": 0})
    if not bon:
        raise HTTPException(status_code=404, detail='Bon sementara tidak ditemukan')
    items_list = [item.model_dump() for item in data.items]
    total_realisasi = sum(item['total'] for item in items_list)
    sisa = bon['jumlah'] - total_realisasi
    doc = {
        "id": str(uuid.uuid4()), "bon_sementara_id": data.bon_sementara_id, "no_bon_ref": bon['no_bon'],
        "user_id": user['id'], "user_name": user['name'], "periode": data.periode,
        "items": items_list, "total_realisasi": total_realisasi, "bon_sementara_amount": bon['jumlah'],
        "sisa_bon": sisa, "bukti_transfer": data.bukti_transfer,
        "status": "pending", "declined_by": None, "decline_reason": None,
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.realisasi_bon.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.put("/realisasi/{real_id}/approve")
async def approve_realisasi(real_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    r = await db.realisasi_bon.find_one({"id": real_id}, {"_id": 0})
    if not r:
        raise HTTPException(status_code=404)
    
    role = user['role']
    current_status = r['status']
    
    # Check if user can approve based on status flow
    flow = REALISASI_STATUS_FLOW.get(current_status)
    if not flow or flow['approver'] != role:
        raise HTTPException(status_code=400, detail=f'Tidak dapat disetujui. Status saat ini: {current_status}, Role: {role}')
    
    new_status = flow['next']
    
    # Track approval history
    approval_history = r.get('approval_history', [])
    approval_history.append({
        "action": "approved",
        "by": user['name'],
        "role": role,
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    await db.realisasi_bon.update_one(
        {"id": real_id}, 
        {"$set": {
            "status": new_status,
            "approval_history": approval_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    status_labels = {
        "approved_hrga": "HRGA",
        "approved_direktur": "Direktur",
        "approved_finance": "Finance (Final)"
    }
    
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), 
        "user_id": r['user_id'], 
        "message": f"Realisasi Bon '{r['no_bon_ref']}' disetujui oleh {status_labels.get(new_status, role)}", 
        "type": "approved", 
        "bon_id": real_id, 
        "is_read": False, 
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Disetujui", "status": new_status}

@api_router.put("/realisasi/{real_id}/decline")
async def decline_realisasi(real_id: str, data: BonAction, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    r = await db.realisasi_bon.find_one({"id": real_id}, {"_id": 0})
    if not r: 
        raise HTTPException(status_code=404)
    
    role = user['role']
    current_status = r['status']
    
    # Check if this role can decline based on current status
    flow = REALISASI_STATUS_FLOW.get(current_status)
    if not flow or flow['approver'] != role:
        raise HTTPException(status_code=403, detail='Tidak dapat menolak pengajuan ini')
    
    # Track approval history
    approval_history = r.get('approval_history', [])
    approval_history.append({
        "action": "declined",
        "by": user['name'],
        "role": role,
        "reason": data.reason or "-",
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    await db.realisasi_bon.update_one(
        {"id": real_id}, 
        {"$set": {
            "status": "declined", 
            "declined_by": role, 
            "decline_reason": data.reason or "-",
            "approval_history": approval_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    role_labels = {
        "hrga": "HRGA",
        "direktur": "Direktur",
        "finance": "Finance"
    }
    
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), 
        "user_id": r['user_id'], 
        "message": f"Realisasi Bon '{r['no_bon_ref']}' ditolak oleh {role_labels.get(role, role)}: {data.reason or '-'}", 
        "type": "declined", 
        "bon_id": real_id, 
        "is_read": False, 
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Ditolak"}

@api_router.put("/realisasi/{real_id}/resubmit")
async def resubmit_realisasi(real_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    if user['role'] != 'pegawai': raise HTTPException(status_code=403)
    r = await db.realisasi_bon.find_one({"id": real_id, "user_id": user['id']}, {"_id": 0})
    if not r or r['status'] != 'declined': raise HTTPException(status_code=400)
    await db.realisasi_bon.update_one({"id": real_id}, {"$set": {"status": "pending", "declined_by": None, "decline_reason": None, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Diajukan ulang"}

@api_router.get("/realisasi/{real_id}/pdf")
async def pdf_realisasi(real_id: str, authorization: str = Header(None)):
    await get_current_user(authorization)
    r = await db.realisasi_bon.find_one({"id": real_id}, {"_id": 0})
    if not r: raise HTTPException(status_code=404)
    if r['status'] != 'approved_finance': raise HTTPException(status_code=400)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'REALISASI BON SEMENTARA', ln=True, align='C')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, f'Tgl Terbit: {r["created_at"][:10]}', ln=True, align='C')
    pdf.ln(5)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 8, 'REKAP REALISASI BON SEMENTARA', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(50, 7, 'Referensi No. Bon Sementara')
    pdf.cell(5, 7, ':')
    pdf.cell(0, 7, r.get('no_bon_ref', '-'), ln=True)
    pdf.cell(50, 7, 'Periode')
    pdf.cell(5, 7, ':')
    pdf.cell(0, 7, r.get('periode', '-'), ln=True)
    pdf.ln(3)
    # Table header
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(10, 8, 'NO', 1, 0, 'C', True)
    pdf.cell(25, 8, 'Tanggal', 1, 0, 'C', True)
    pdf.cell(65, 8, 'Uraian', 1, 0, 'C', True)
    pdf.cell(15, 8, 'Qty', 1, 0, 'C', True)
    pdf.cell(35, 8, 'Harga/Unit', 1, 0, 'C', True)
    pdf.cell(35, 8, 'Total', 1, 1, 'C', True)
    pdf.set_font('Helvetica', '', 9)
    for i, item in enumerate(r.get('items', []), 1):
        pdf.cell(10, 7, str(i), 1, 0, 'C')
        pdf.cell(25, 7, str(item.get('tanggal', '')), 1, 0, 'C')
        pdf.cell(65, 7, str(item.get('uraian', '')), 1)
        pdf.cell(15, 7, str(item.get('quantity', 1)), 1, 0, 'C')
        pdf.cell(35, 7, f'Rp {item.get("harga_per_unit", 0):,.0f}', 1, 0, 'R')
        pdf.cell(35, 7, f'Rp {item.get("total", 0):,.0f}', 1, 1, 'R')
    pdf.ln(3)
    pdf.set_font('Helvetica', 'B', 10)
    for label, val in [("TOTAL", r.get('total_realisasi', 0)), ("BON SEMENTARA", r.get('bon_sementara_amount', 0)), ("SISA BON SEMENTARA", r.get('sisa_bon', 0))]:
        pdf.cell(115, 7, label)
        pdf.cell(0, 7, f'Rp {val:,.0f}', ln=True)
    pdf.ln(15)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(60, 7, 'Dibuat oleh,', align='C')
    pdf.cell(60, 7, 'Diketahui oleh,', align='C')
    pdf.cell(60, 7, 'Diverifikasi oleh,', align='C', ln=True)
    pdf.ln(18)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(60, 7, r['user_name'], align='C')
    pdf.cell(60, 7, 'Atasan Langsung', align='C')
    pdf.cell(60, 7, 'Finance', align='C')
    return Response(content=bytes(pdf.output()), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=realisasi-{r['no_bon_ref'].replace('/', '-')}.pdf"})

# --- Notifications ---
@api_router.get("/notifications")
async def get_notifications(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    return await db.notifications.find({"user_id": user['id']}, {"_id": 0}).sort("created_at", -1).to_list(50)

@api_router.put("/notifications/read-all")
async def mark_all_read(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    await db.notifications.update_many({"user_id": user['id'], "is_read": False}, {"$set": {"is_read": True}})
    return {"message": "OK"}

# --- Pengaduan ---
@api_router.get("/pengaduan")
async def get_pengaduan(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    q = {"user_id": user['id']} if user['role'] == 'pegawai' else {}
    return await db.pengaduan.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)

@api_router.post("/pengaduan")
async def create_pengaduan(data: PengaduanCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    item = {"id": str(uuid.uuid4()), "user_id": user['id'], "user_name": user['name'], "judul": data.judul, "deskripsi": data.deskripsi, "kategori": data.kategori, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat()}
    await db.pengaduan.insert_one(item)
    item.pop('_id', None)
    return item

# --- Cuti ---
@api_router.get("/cuti")
async def get_cuti(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    q = {"user_id": user['id']} if user['role'] == 'pegawai' else {}
    return await db.cuti.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)

@api_router.post("/cuti")
async def create_cuti(data: CutiCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    item = {"id": str(uuid.uuid4()), "user_id": user['id'], "user_name": user['name'], "tanggal_mulai": data.tanggal_mulai, "tanggal_selesai": data.tanggal_selesai, "jenis": data.jenis, "alasan": data.alasan, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat()}
    await db.cuti.insert_one(item)
    item.pop('_id', None)
    return item

# --- Inventaris ---
@api_router.get("/inventaris")
async def get_inventaris(authorization: str = Header(None)):
    await get_current_user(authorization)
    return await db.inventaris.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api_router.post("/inventaris")
async def create_inventaris(data: InventarisCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    item = {"id": str(uuid.uuid4()), "nama": data.nama, "kategori": data.kategori, "jumlah": data.jumlah, "kondisi": data.kondisi, "lokasi": data.lokasi, "created_by": user['name'], "created_at": datetime.now(timezone.utc).isoformat()}
    await db.inventaris.insert_one(item)
    item.pop('_id', None)
    return item

app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','), allow_methods=["*"], allow_headers=["*"])

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    count = await db.users.count_documents({})
    if count == 0:
        users = [
            {'id': str(uuid.uuid4()), 'name': 'Ahmad Pegawai', 'email': 'pegawai@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'pegawai'},
            {'id': str(uuid.uuid4()), 'name': 'Budi Atasan', 'email': 'atasan@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'atasan'},
            {'id': str(uuid.uuid4()), 'name': 'Dewi HRGA', 'email': 'hrga@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'hrga'},
            {'id': str(uuid.uuid4()), 'name': 'Eko Direktur', 'email': 'direktur@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'direktur'},
            {'id': str(uuid.uuid4()), 'name': 'Citra Finance', 'email': 'finance@kantor.com', 'password': pwd_context.hash('password123'), 'role': 'finance'},
        ]
        await db.users.insert_many(users)
        logger.info("Seed data created with 5 users")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
